#!/usr/bin/env python3
"""Audit GSE243639 cell-ID matching across expression, cell map, and annotation workbook."""

from __future__ import annotations

import argparse
import csv
import gzip
import logging
import re
from pathlib import Path
from typing import TextIO


CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def open_text(path: Path) -> TextIO:
    if path.name.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8", newline="")
    return path.open("r", encoding="utf-8", newline="")


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX cell-ID audit. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def normalize_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_cell_id(cell_id: str) -> tuple[str, str]:
    value = cell_id.strip()
    if "_" in value:
        sample_id, barcode = value.split("_", 1)
        return sample_id, barcode
    return "", value


def remove_trailing_gem_suffix(value: str) -> str:
    return re.sub(r"\.\d+$", "", value.strip())


def normalized_punctuation(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value).upper()


def cell_id_variants(cell_id: str) -> dict[str, str]:
    sample_id, barcode = split_cell_id(cell_id)
    core_with_suffix = barcode.strip()
    core_without_suffix = remove_trailing_gem_suffix(core_with_suffix)
    full_without_suffix = f"{sample_id}_{core_without_suffix}" if sample_id else core_without_suffix
    return {
        "original": cell_id.strip(),
        "sample_id": sample_id,
        "barcode_core": core_without_suffix,
        "without_sample_prefix": core_with_suffix,
        "without_trailing_suffix": full_without_suffix,
        "punctuation_normalized": normalized_punctuation(full_without_suffix),
    }


def read_expression_cell_ids(path: Path) -> set[str]:
    ids: set[str] = set()
    with open_text(path) as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id:
                ids.add(cell_id)
    return ids


def read_cell_sample_ids(path: Path) -> set[str]:
    ids: set[str] = set()
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id:
                ids.add(cell_id)
    return ids


def role_score(column: str) -> int:
    lowered = column.lower().replace(" ", "_")
    return sum(1 for term in CELL_ID_TERMS if term in lowered)


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:20]):
        score = sum(role_score(value) for value in row) + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_cell_column(header: list[str]) -> str:
    scored = sorted(((role_score(column), column) for column in header if column), reverse=True)
    return scored[0][1] if scored and scored[0][0] > 0 else (header[0] if header else "")


def read_workbook_cell_ids(path: Path) -> tuple[set[str], str, str]:
    workbook = load_workbook_readonly(path)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    preview = [[normalize_text(value) for value in row] for row in sheet.iter_rows(max_row=30, values_only=True)]
    header_index = infer_header_row(preview)
    header = preview[header_index]
    cell_column = choose_cell_column(header)
    column_index = header.index(cell_column)
    ids: set[str] = set()
    for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
        values = [normalize_text(value) for value in row]
        if column_index < len(values) and values[column_index]:
            ids.add(values[column_index])
    return ids, sheet_name, cell_column


def variant_set(ids: set[str], variant: str) -> set[str]:
    return {cell_id_variants(cell_id)[variant] for cell_id in ids if cell_id}


def recommended_rule(expression_ids: set[str], workbook_ids: set[str]) -> tuple[str, int]:
    direct = len(expression_ids & workbook_ids)
    suffix = len(variant_set(expression_ids, "without_trailing_suffix") & variant_set(workbook_ids, "without_trailing_suffix"))
    core = len(variant_set(expression_ids, "barcode_core") & variant_set(workbook_ids, "barcode_core"))
    punct = len(variant_set(expression_ids, "punctuation_normalized") & variant_set(workbook_ids, "punctuation_normalized"))
    options = [
        ("direct_original_cell_id", direct),
        ("remove_trailing_dot_suffix", suffix),
        ("match_barcode_core_after_sample_prefix_removal", core),
        ("normalize_punctuation_and_suffix", punct),
    ]
    return max(options, key=lambda item: item[1])


def write_audit(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["metric", "value", "recommended_normalization_rule", "notes"],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_preview(path: Path, expression_ids: set[str], workbook_ids: set[str], limit: int = 20) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    expression_preview = sorted(expression_ids)[:limit]
    workbook_preview = sorted(workbook_ids)[:limit]
    fields = ["row_index", "expression_cell_id", "expression_barcode_core", "workbook_cell_id", "workbook_barcode_core"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        for index in range(max(len(expression_preview), len(workbook_preview))):
            expr = expression_preview[index] if index < len(expression_preview) else ""
            anno = workbook_preview[index] if index < len(workbook_preview) else ""
            writer.writerow(
                {
                    "row_index": str(index + 1),
                    "expression_cell_id": expr,
                    "expression_barcode_core": cell_id_variants(expr)["barcode_core"] if expr else "",
                    "workbook_cell_id": anno,
                    "workbook_barcode_core": cell_id_variants(anno)["barcode_core"] if anno else "",
                }
            )


def build_audit(args: argparse.Namespace) -> tuple[list[dict[str, str]], set[str], set[str]]:
    expression_ids = read_expression_cell_ids(args.expression)
    map_ids = read_cell_sample_ids(args.cell_sample_map)
    workbook_ids, sheet_name, cell_column = read_workbook_cell_ids(args.xlsx)
    rule, rule_overlap = recommended_rule(expression_ids, workbook_ids)
    metrics = {
        "unique_expression_cell_ids": len(expression_ids),
        "unique_cell_sample_map_cell_ids": len(map_ids),
        "unique_workbook_cell_ids": len(workbook_ids),
        "direct_overlap_count": len(expression_ids & workbook_ids),
        "expression_vs_cell_sample_map_overlap": len(expression_ids & map_ids),
        "overlap_after_removing_sample_prefix": len(variant_set(expression_ids, "without_sample_prefix") & variant_set(workbook_ids, "without_sample_prefix")),
        "overlap_after_adding_sample_prefix": len(variant_set(expression_ids, "without_sample_prefix") & variant_set(workbook_ids, "original")),
        "overlap_after_removing_trailing_dot_suffix": len(variant_set(expression_ids, "without_trailing_suffix") & variant_set(workbook_ids, "without_trailing_suffix")),
        "overlap_after_normalizing_punctuation": len(variant_set(expression_ids, "punctuation_normalized") & variant_set(workbook_ids, "punctuation_normalized")),
        "recommended_rule_overlap": rule_overlap,
    }
    rows = [
        {
            "metric": metric,
            "value": str(value),
            "recommended_normalization_rule": rule,
            "notes": f"Workbook sheet={sheet_name}; workbook cell column={cell_column}",
        }
        for metric, value in metrics.items()
    ]
    return rows, expression_ids, workbook_ids


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit GSE243639 cell ID matching.")
    parser.add_argument("--expression", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_sparse_gene_panel_expression.tsv.gz"))
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("results/tables/phase18_gse243639_cell_id_matching_audit.tsv"))
    parser.add_argument("--preview-output", type=Path, default=Path("results/reports/phase18_gse243639_cell_id_matching_preview.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/90_audit_gse243639_cell_id_matching.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    rows, expression_ids, workbook_ids = build_audit(args)
    write_audit(args.output, rows)
    write_preview(args.preview_output, expression_ids, workbook_ids)
    logging.info("GSE243639 cell-ID audit rows: %d", len(rows))
    print(f"Wrote {args.output}")
    print(f"Wrote {args.preview_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
