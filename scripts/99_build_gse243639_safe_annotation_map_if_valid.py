#!/usr/bin/env python3
"""Build a GSE243639 annotation map only when Phase 19 linkage is safe."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path


SAFE_DECISIONS = {
    "direct_id_linkage_safe",
    "normalized_id_linkage_safe",
    "row_order_linkage_safe_with_caution",
}
BIOLOGICAL_LABEL_TERMS = [
    "astrocyte",
    "microglia",
    "oligodendrocyte",
    "neuron",
    "opc",
    "endothelial",
    "pericyte",
    "dopaminergic",
    "excitatory",
    "inhibitory",
]
CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]
SAMPLE_ID_TERMS = ["sample", "donor", "subject"]
ANNOTATION_TERMS = ["celltype", "cell_type", "annotation", "cluster", "subcluster", "class", "type"]
MANDATORY_COLUMNS = [
    "cell_id_expression",
    "cell_id_annotation_original",
    "cell_id_annotation_normalized",
    "normalized_cell_id",
    "barcode_core",
    "sample_id",
    "cell_type",
    "cluster_id",
    "annotation_source_sheet",
    "annotation_column_used",
    "match_status",
    "normalization_rule",
]


def read_tsv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_blocked_report(path: Path, decision: str) -> None:
    lines = [
        "# Phase 19 Annotation Linkage Blocked",
        "",
        f"- Decision: `{decision or 'not_available'}`",
        "- Action: no safe cell annotation map was generated.",
        "",
        "Cell-type-aware GSE243639 validation is not currently supported from the available workbook because cell IDs cannot be safely linked to expression cells.",
        "Use the Phase 16 global sample-level PD extension unless a new audited annotation source becomes available.",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_") or "missing"


def split_id(value: str) -> tuple[str, str]:
    stripped = value.strip().strip("\"'")
    if "_" in stripped:
        sample_id, barcode = stripped.split("_", 1)
        return sample_id, barcode
    return "", stripped


def remove_trailing(value: str) -> str:
    return re.sub(r"([.-])\d+$", "", value.strip())


def collapse_punctuation(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value.strip()).upper()


def normalize_by_rule(value: str, rule: str) -> str:
    if rule in {"", "raw_id", "direct_id_linkage_safe"}:
        return value.strip()
    if rule == "lowercase":
        return value.lower()
    if rule == "remove_quotes":
        return value.strip("\"'")
    if rule == "replace_dash_with_dot":
        return value.replace("-", ".")
    if rule == "replace_dot_with_dash":
        return value.replace(".", "-")
    if rule in {"remove_trailing_dot_or_dash_one", "remove_trailing_dot_suffix"}:
        return remove_trailing(value)
    if rule == "keep_barcode_only":
        return remove_trailing(split_id(value)[1])
    if rule == "keep_sample_prefix_only":
        return split_id(value)[0]
    if rule == "remove_sample_prefix":
        return split_id(value)[1]
    if rule == "collapse_punctuation":
        return collapse_punctuation(value)
    return value.strip()


def barcode_core(value: str) -> str:
    return remove_trailing(split_id(value)[1])


def role(column: str) -> str:
    lowered = column.lower().replace(" ", "_")
    if any(term in lowered for term in ANNOTATION_TERMS):
        return "annotation"
    if any(term in lowered for term in CELL_ID_TERMS):
        return "cell_id"
    if any(term in lowered for term in SAMPLE_ID_TERMS):
        return "sample_id"
    return "unmapped"


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        score = sum(1 for value in row if role(value) != "unmapped") + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_column(columns: list[str], wanted: str) -> str:
    candidates = [column for column in columns if role(column) == wanted]
    return candidates[0] if candidates else ""


def biological_label_score(values: list[str]) -> int:
    lowered_values = [value.lower() for value in values if value]
    return sum(any(term in value for term in BIOLOGICAL_LABEL_TERMS) for value in lowered_values)


def numeric_fraction(values: list[str]) -> float:
    observed = [value for value in values if value]
    if not observed:
        return 1.0
    numeric = sum(1 for value in observed if re.fullmatch(r"\d+(\.0)?", value))
    return numeric / len(observed)


def choose_annotation_column(columns: list[str], rows: list[dict[str, str]]) -> tuple[str, str]:
    candidates = [column for column in columns if role(column) == "annotation"]
    if not candidates:
        return "", "none"
    scored = []
    for column in candidates:
        values = [row.get(column, "") for row in rows[:300]]
        bio_score = biological_label_score(values)
        num_fraction = numeric_fraction(values)
        if bio_score > 0:
            priority = 0
            kind = "biological_cell_type"
        elif "cluster" in column.lower() or num_fraction >= 0.80:
            priority = 2
            kind = "cluster"
        else:
            priority = 1
            kind = "annotation"
        scored.append((priority, column, kind))
    scored.sort()
    return scored[0][1], scored[0][2]


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for safe annotation map creation. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def read_workbook_rows(xlsx: Path) -> tuple[str, list[str], list[dict[str, str]]]:
    workbook = load_workbook_readonly(xlsx)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    preview = [[clean(value) for value in row] for row in sheet.iter_rows(max_row=40, values_only=True)]
    if not preview:
        return sheet_name, [], []
    header_index = infer_header_row(preview)
    header = preview[header_index]
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
        values = [clean(value) for value in row]
        if any(values):
            rows.append({column: value for column, value in zip(header, values, strict=False) if column})
    return sheet_name, header, rows


def read_cell_sample_map(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            sample_id = row.get("sample_id", "")
            if cell_id:
                mapping[cell_id] = sample_id or split_id(cell_id)[0]
    return mapping


def expression_lookup(cell_sample_map: Path, rule: str) -> tuple[dict[str, list[tuple[str, str]]], dict[str, str]]:
    lookup: dict[str, list[tuple[str, str]]] = defaultdict(list)
    expression_samples = read_cell_sample_map(cell_sample_map)
    for expression_id, sample_id in expression_samples.items():
        lookup[normalize_by_rule(expression_id, rule)].append((expression_id, sample_id))
    return lookup, expression_samples


def decision_info(path: Path) -> tuple[str, str, bool]:
    rows = read_tsv(path)
    if not rows:
        return "", "raw_id", False
    row = rows[0]
    decision = row.get("decision_category", "")
    rule = row.get("direct_or_normalized_best_rule", "") or "raw_id"
    safe = row.get("safe_to_build_annotation_map") == "true" and decision in SAFE_DECISIONS
    return decision, rule, safe


def build_safe_rows(decision_path: Path, cell_sample_map: Path, xlsx: Path) -> tuple[str, bool, list[dict[str, str]]]:
    decision, rule, safe = decision_info(decision_path)
    if not safe:
        return decision, False, []
    sheet_name, columns, workbook_rows = read_workbook_rows(xlsx)
    cell_column = choose_column(columns, "cell_id") or (columns[0] if columns else "")
    sample_column = choose_column(columns, "sample_id")
    annotation_column, annotation_kind = choose_annotation_column(columns, workbook_rows)
    lookup, expression_samples = expression_lookup(cell_sample_map, rule)
    output_rows: list[dict[str, str]] = []
    used_expression_ids: set[str] = set()
    for row in workbook_rows:
        annotation_original = row.get(cell_column, "")
        if not annotation_original:
            continue
        normalized_annotation = normalize_by_rule(annotation_original, rule)
        matches = lookup.get(normalized_annotation, [])
        if len(matches) != 1:
            continue
        expression_id, sample_id = matches[0]
        if expression_id in used_expression_ids:
            continue
        used_expression_ids.add(expression_id)
        raw_label = row.get(annotation_column, "") if annotation_column else ""
        cluster_id = raw_label if annotation_kind == "cluster" else ""
        cell_type = f"cluster_{safe_name(raw_label)}" if annotation_kind == "cluster" else (raw_label or "unannotated")
        output_rows.append(
            {
                "cell_id_expression": expression_id,
                "cell_id_annotation_original": annotation_original,
                "cell_id_annotation_normalized": normalized_annotation,
                "normalized_cell_id": normalize_by_rule(expression_id, rule),
                "barcode_core": barcode_core(expression_id),
                "sample_id": sample_id or row.get(sample_column, "") or split_id(expression_id)[0],
                "cell_type": cell_type,
                "cluster_id": cluster_id,
                "annotation_source_sheet": sheet_name,
                "annotation_column_used": annotation_column,
                "match_status": "matched",
                "normalization_rule": rule,
            }
        )
    # Preserve expression cells without workbook rows as explicit unmatched records only for auditability.
    for expression_id, sample_id in expression_samples.items():
        if expression_id in used_expression_ids:
            continue
        output_rows.append(
            {
                "cell_id_expression": expression_id,
                "cell_id_annotation_original": "",
                "cell_id_annotation_normalized": "",
                "normalized_cell_id": normalize_by_rule(expression_id, rule),
                "barcode_core": barcode_core(expression_id),
                "sample_id": sample_id,
                "cell_type": "unmatched",
                "cluster_id": "",
                "annotation_source_sheet": sheet_name,
                "annotation_column_used": annotation_column,
                "match_status": "unmatched",
                "normalization_rule": rule,
            }
        )
    return decision, True, output_rows


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANDATORY_COLUMNS, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build safe GSE243639 annotation map only after Phase 19 approval.")
    parser.add_argument("--decision", type=Path, default=Path("results/tables/phase19_gse243639_annotation_linkage_decision.tsv"))
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--existing-annotation-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_annotation_map.tsv"))
    parser.add_argument("--output", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_safe_cell_annotation_map.tsv"))
    parser.add_argument("--blocked-output", type=Path, default=Path("results/reports/phase19_annotation_linkage_blocked.md"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    decision, safe, rows = build_safe_rows(args.decision, args.cell_sample_map, args.xlsx)
    if not safe:
        write_blocked_report(args.blocked_output, decision)
        print(f"Blocked annotation map creation; wrote {args.blocked_output}")
        return 0
    write_rows(args.output, rows)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
