#!/usr/bin/env python3
"""Inspect GSE243639 UMAP coordinate workbook as an annotation table only."""

from __future__ import annotations

import argparse
import csv
import logging
from pathlib import Path


CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei"]
SAMPLE_ID_TERMS = ["sample", "donor", "subject"]
ANNOTATION_TERMS = ["celltype", "cell_type", "annotation", "cluster", "subcluster", "class", "type"]
UMAP_TERMS = ["umap", "x_umap", "umap_1", "umap1", "umap2", "umap_2"]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX annotation inspection. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def likely_role(column: str) -> str:
    lowered = column.lower().replace(" ", "_")
    if any(term in lowered for term in UMAP_TERMS):
        return "candidate_umap_coordinate"
    if any(term in lowered for term in ANNOTATION_TERMS):
        return "candidate_celltype_or_cluster_annotation"
    if any(term in lowered for term in CELL_ID_TERMS):
        return "candidate_cell_id"
    if any(term in lowered for term in SAMPLE_ID_TERMS):
        return "candidate_sample_id"
    return "unmapped"


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows):
        score = sum(1 for value in row if likely_role(value) != "unmapped") + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def inspect_workbook(path: Path, preview_limit: int = 20) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook = load_workbook_readonly(path)
    audit_rows: list[dict[str, str]] = []
    preview_rows: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        raw_rows = [[normalize(value) for value in row] for row in worksheet.iter_rows(max_row=preview_limit + 10, values_only=True)]
        if not raw_rows:
            continue
        header_index = infer_header_row(raw_rows)
        header = raw_rows[header_index]
        for position, column in enumerate(header, start=1):
            if not column:
                continue
            audit_rows.append(
                {
                    "sheet_name": sheet_name,
                    "header_line": str(header_index + 1),
                    "column_index": str(position),
                    "column_name": column,
                    "likely_role": likely_role(column),
                    "note": "Workbook inspected as annotation/coordinate table only; no UMAP computation performed.",
                }
            )
        for row_index, row in enumerate(raw_rows[header_index + 1 : header_index + 1 + preview_limit], start=1):
            if not any(row):
                continue
            preview = {"sheet_name": sheet_name, "preview_row_index": str(row_index)}
            for column, value in zip(header, row, strict=False):
                if column:
                    preview[column] = value
            preview_rows.append(preview)
    return audit_rows, preview_rows


def write_tsv(path: Path, rows: list[dict[str, str]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = sorted({field for row in rows for field in row}) if rows else ["status"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        if rows:
            writer.writerows(rows)
        else:
            writer.writerow({"status": "no_rows"})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect GSE243639 UMAP annotation workbook safely.")
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("results/reports/phase17_gse243639_umap_annotation_audit.tsv"))
    parser.add_argument("--preview-output", type=Path, default=Path("results/reports/phase17_gse243639_umap_annotation_preview.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/83_inspect_gse243639_umap_annotations.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    audit_rows, preview_rows = inspect_workbook(args.xlsx)
    write_tsv(args.output, audit_rows, ["sheet_name", "header_line", "column_index", "column_name", "likely_role", "note"])
    write_tsv(args.preview_output, preview_rows)
    logging.info("Inspected GSE243639 annotation workbook columns: %d", len(audit_rows))
    print(f"Wrote {args.output}")
    print(f"Wrote {args.preview_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
