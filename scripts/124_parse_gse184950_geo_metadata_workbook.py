#!/usr/bin/env python3
"""Parse the small GSE184950 GEO metadata workbook without opening expression data."""

from __future__ import annotations

import argparse
import csv
import logging
from pathlib import Path


OUTPUT_COLUMNS = [
    "sample_name",
    "title",
    "source_name",
    "organism",
    "tissue",
    "disease_state",
    "donor_id",
    "age",
    "gender",
    "race",
    "ethnicity",
    "pmi_hours",
    "braak_stage",
    "processed_data_file",
    "raw_files",
]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def write_tsv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def norm(value: object) -> str:
    return str(value or "").strip()


def key(value: object) -> str:
    return norm(value).lower().replace("_", " ").replace("-", " ")


def load_workbook_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to parse .xlsx metadata workbooks. Install project dependencies first.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "METADATA TEMPLATE" if "METADATA TEMPLATE" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = [[norm(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def find_sample_header(rows: list[list[str]]) -> int | None:
    samples_section_seen = False
    for index, row in enumerate(rows):
        keys = {key(cell) for cell in row if cell}
        if "samples" in keys:
            samples_section_seen = True
        if "sample name" in keys and (samples_section_seen or "processed data file" in keys or "raw file" in keys):
            return index
    return None


def map_header(header: list[str]) -> dict[str, int]:
    aliases = {
        "sample_name": ["sample name"],
        "title": ["title"],
        "source_name": ["source name"],
        "organism": ["organism", "organism part"],
        "tissue": ["characteristics: tissue", "tissue"],
        "disease_state": ["characteristics: disease state", "disease state"],
        "donor_id": ["characteristics: brain bank donor id", "brain bank donor id", "donor id"],
        "age": ["characteristics: age", "age"],
        "gender": ["characteristics: gender", "gender", "sex"],
        "race": ["characteristics: race", "race"],
        "ethnicity": ["characteristics: ethnicity", "ethnicity"],
        "pmi_hours": ["characteristics: postmortem interval hours", "postmortem interval hours", "pmi hours"],
        "braak_stage": ["characteristics: braak stage", "braak stage"],
        "processed_data_file": ["processed data file"],
        "raw_files": ["raw file", "raw files"],
    }
    lookup = {key(value): index for index, value in enumerate(header) if value}
    mapping: dict[str, int] = {}
    for output, candidates in aliases.items():
        for candidate in candidates:
            if candidate in lookup:
                mapping[output] = lookup[candidate]
                break
    return mapping


def parse_samples(rows: list[list[str]]) -> list[dict[str, str]]:
    header_index = find_sample_header(rows)
    if header_index is None:
        return []
    header = rows[header_index]
    mapping = map_header(header)
    sample_rows: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        if not any(row):
            if sample_rows:
                break
            continue
        sample_name_index = mapping.get("sample_name")
        if sample_name_index is None or sample_name_index >= len(row) or not row[sample_name_index]:
            continue
        parsed = {column: "" for column in OUTPUT_COLUMNS}
        for column, index in mapping.items():
            parsed[column] = row[index] if index < len(row) else ""
        sample_rows.append(parsed)
    return sample_rows


def processed_manifest(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    manifest = []
    for row in rows:
        for file_name in [part.strip() for part in row.get("processed_data_file", "").replace(",", ";").split(";") if part.strip()]:
            manifest.append(
                {
                    "sample_name": row.get("sample_name", ""),
                    "donor_id": row.get("donor_id", ""),
                    "disease_state": row.get("disease_state", ""),
                    "processed_data_file": file_name,
                    "expected_archive_member": file_name,
                    "status": "metadata_declared",
                }
            )
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse GSE184950_add2.xlsx sample metadata.")
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse184950_pd_sn/GSE184950_add2.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("results/tables/phase24_gse184950_sample_metadata.tsv"))
    parser.add_argument("--processed-files-output", type=Path, default=Path("results/tables/phase24_gse184950_processed_file_manifest.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/124_parse_gse184950_geo_metadata_workbook.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    rows = parse_samples(load_workbook_rows(args.xlsx))
    write_tsv(args.output, rows, OUTPUT_COLUMNS)
    manifest = processed_manifest(rows)
    write_tsv(args.processed_files_output, manifest, ["sample_name", "donor_id", "disease_state", "processed_data_file", "expected_archive_member", "status"])
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.get("disease_state", "missing")] = counts.get(row.get("disease_state", "missing"), 0) + 1
    logging.info("Parsed GSE184950 workbook samples=%d disease_counts=%s processed_files=%d", len(rows), counts, len(manifest))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
