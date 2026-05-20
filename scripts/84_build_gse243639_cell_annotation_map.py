#!/usr/bin/env python3
"""Build a repaired GSE243639 cell annotation map with robust cell-ID matching."""

from __future__ import annotations

import argparse
import csv
import logging
import re
from collections import Counter, defaultdict
from pathlib import Path


BIOLOGICAL_LABEL_TERMS = [
    "astrocyte",
    "microglia",
    "oligodendrocyte",
    "neuron",
    "opc",
    "endothelial",
    "pericyte",
    "dopaminergic",
    "excitatory",
    "inhibitory",
]
CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]
SAMPLE_ID_TERMS = ["sample", "donor", "subject"]
ANNOTATION_TERMS = ["celltype", "cell_type", "annotation", "cluster", "subcluster", "class", "type"]
UMAP_TERMS = ["umap", "x_umap", "umap_1", "umap1", "umap2", "umap_2"]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX annotation mapping. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_") or "missing"


def split_cell_id(cell_id: str) -> tuple[str, str]:
    value = cell_id.strip()
    if "_" in value:
        sample_id, barcode = value.split("_", 1)
        return sample_id, barcode
    return "", value


def remove_trailing_gem_suffix(value: str) -> str:
    return re.sub(r"\.\d+$", "", value.strip())


def normalized_punctuation(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value).upper()


def cell_id_parts(cell_id: str, fallback_sample: str = "") -> dict[str, str]:
    sample_id, barcode = split_cell_id(cell_id)
    sample_id = sample_id or fallback_sample
    barcode_core = remove_trailing_gem_suffix(barcode)
    normalized = f"{sample_id}_{barcode_core}" if sample_id else barcode_core
    return {
        "original": cell_id.strip(),
        "sample_id": sample_id,
        "normalized_cell_id": normalized,
        "barcode_core": barcode_core,
        "punctuation_normalized": normalized_punctuation(normalized),
        "barcode_punctuation_normalized": normalized_punctuation(barcode_core),
    }


def role(column: str) -> str:
    lowered = column.lower().replace(" ", "_")
    if any(term in lowered for term in UMAP_TERMS):
        return "umap"
    if any(term in lowered for term in ANNOTATION_TERMS):
        return "annotation"
    if any(term in lowered for term in CELL_ID_TERMS):
        return "cell_id"
    if any(term in lowered for term in SAMPLE_ID_TERMS):
        return "sample_id"
    return "unmapped"


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:25]):
        score = sum(1 for value in row if role(value) != "unmapped") + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_cell_id_column(columns: list[str]) -> str:
    candidates = [column for column in columns if role(column) == "cell_id"]
    return candidates[0] if candidates else (columns[0] if columns else "")


def choose_sample_id_column(columns: list[str]) -> str:
    candidates = [column for column in columns if role(column) == "sample_id"]
    return candidates[0] if candidates else ""


def biological_label_score(values: list[str]) -> int:
    lowered_values = [value.lower() for value in values if value]
    return sum(any(term in value for term in BIOLOGICAL_LABEL_TERMS) for value in lowered_values)


def numeric_fraction(values: list[str]) -> float:
    observed = [value for value in values if value]
    if not observed:
        return 1.0
    numeric = sum(1 for value in observed if re.fullmatch(r"\d+(\.0)?", value))
    return numeric / len(observed)


def choose_annotation_column(columns: list[str], preview_rows: list[dict[str, str]]) -> tuple[str, str, str, list[dict[str, str]]]:
    candidates = [column for column in columns if role(column) == "annotation"]
    candidate_rows: list[dict[str, str]] = []
    for column in candidates:
        values = [row.get(column, "") for row in preview_rows[:200]]
        bio_score = biological_label_score(values)
        num_fraction = numeric_fraction(values)
        if bio_score > 0:
            confidence = "high"
            annotation_kind = "biological_cell_type"
        elif "cluster" in column.lower() or num_fraction >= 0.80:
            confidence = "low"
            annotation_kind = "cluster"
        else:
            confidence = "medium"
            annotation_kind = "annotation"
        candidate_rows.append(
            {
                "column_name": column,
                "annotation_kind": annotation_kind,
                "biological_label_hits": str(bio_score),
                "numeric_fraction": f"{num_fraction:.4f}",
                "confidence": confidence,
            }
        )
    high = [row for row in candidate_rows if row["confidence"] == "high"]
    medium = [row for row in candidate_rows if row["confidence"] == "medium"]
    chosen = (high or medium or candidate_rows or [{"column_name": "", "annotation_kind": "none", "confidence": "low"}])[0]
    return chosen["column_name"], chosen["annotation_kind"], chosen["confidence"], candidate_rows


def read_cell_sample_map(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            sample_id = row.get("sample_id", "")
            if cell_id and sample_id:
                mapping[cell_id] = sample_id
    return mapping


def read_workbook_rows(xlsx: Path) -> tuple[str, list[str], list[dict[str, str]]]:
    workbook = load_workbook_readonly(xlsx)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    preview = [[normalize(value) for value in row] for row in sheet.iter_rows(max_row=40, values_only=True)]
    header_index = infer_header_row(preview)
    header = preview[header_index]
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
        values = [normalize(value) for value in row]
        if not any(values):
            continue
        rows.append({column: value for column, value in zip(header, values, strict=False) if column})
    return sheet_name, header, rows


def build_annotation_lookup(
    workbook_rows: list[dict[str, str]],
    cell_column: str,
    sample_column: str,
    annotation_column: str,
    annotation_kind: str,
    confidence: str,
    sheet_name: str,
) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    keyed: dict[str, dict[str, str]] = {}
    barcode_keyed: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in workbook_rows:
        original = row.get(cell_column, "")
        if not original:
            continue
        fallback_sample = row.get(sample_column, "") if sample_column else ""
        parts = cell_id_parts(original, fallback_sample=fallback_sample)
        raw_label = row.get(annotation_column, "") if annotation_column else ""
        cluster_id = raw_label if annotation_kind == "cluster" else ""
        cell_type = f"cluster_{safe_name(raw_label)}" if annotation_kind == "cluster" else (raw_label or "unannotated")
        record = {
            "cell_id_annotation_original": original,
            "normalized_cell_id": parts["normalized_cell_id"],
            "barcode_core": parts["barcode_core"],
            "sample_id": parts["sample_id"],
            "cell_type": cell_type,
            "cluster_id": cluster_id,
            "annotation_source_sheet": sheet_name,
            "annotation_column_used": annotation_column,
            "biological_celltype_confidence": confidence,
        }
        keyed[parts["normalized_cell_id"]] = record
        keyed[parts["punctuation_normalized"]] = record
        barcode_keyed[parts["barcode_core"]].append(record)
        barcode_keyed[parts["barcode_punctuation_normalized"]].append(record)
    return keyed, barcode_keyed


def match_annotation(
    expression_cell_id: str,
    expression_sample_id: str,
    keyed: dict[str, dict[str, str]],
    barcode_keyed: dict[str, list[dict[str, str]]],
) -> tuple[dict[str, str] | None, str]:
    parts = cell_id_parts(expression_cell_id, fallback_sample=expression_sample_id)
    for key in [parts["normalized_cell_id"], parts["punctuation_normalized"]]:
        if key in keyed:
            return keyed[key], "matched_normalized_cell_id"
    barcode_matches = barcode_keyed.get(parts["barcode_core"], []) or barcode_keyed.get(parts["barcode_punctuation_normalized"], [])
    if barcode_matches:
        sample_matches = [record for record in barcode_matches if record.get("sample_id") in {"", expression_sample_id}]
        if len(sample_matches) == 1:
            return sample_matches[0], "matched_barcode_core"
        if len(barcode_matches) == 1:
            return barcode_matches[0], "matched_unique_barcode_core"
    return None, "unmatched"


def build_annotation_map(
    xlsx: Path,
    cell_sample_map: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    expression_cells = read_cell_sample_map(cell_sample_map)
    sheet_name, columns, workbook_rows = read_workbook_rows(xlsx)
    cell_column = choose_cell_id_column(columns)
    sample_column = choose_sample_id_column(columns)
    annotation_column, annotation_kind, confidence, candidate_rows = choose_annotation_column(columns, workbook_rows)
    keyed, barcode_keyed = build_annotation_lookup(
        workbook_rows,
        cell_column,
        sample_column,
        annotation_column,
        annotation_kind,
        confidence,
        sheet_name,
    )
    output_rows: list[dict[str, str]] = []
    matched = 0
    counts: Counter[tuple[str, str]] = Counter()
    for expression_cell_id, expression_sample_id in sorted(expression_cells.items()):
        parts = cell_id_parts(expression_cell_id, fallback_sample=expression_sample_id)
        record, status = match_annotation(expression_cell_id, expression_sample_id, keyed, barcode_keyed)
        if record is not None:
            matched += 1
            cell_type = record["cell_type"]
            cluster_id = record["cluster_id"]
            confidence_value = record["biological_celltype_confidence"]
            annotation_original = record["cell_id_annotation_original"]
        else:
            cell_type = "unmatched"
            cluster_id = ""
            confidence_value = "none"
            annotation_original = ""
        output_rows.append(
            {
                "cell_id_expression": expression_cell_id,
                "cell_id_annotation_original": annotation_original,
                "normalized_cell_id": parts["normalized_cell_id"],
                "barcode_core": parts["barcode_core"],
                "sample_id": expression_sample_id,
                "cell_type": cell_type,
                "cluster_id": cluster_id,
                "biological_celltype_confidence": confidence_value,
                "annotation_source_sheet": sheet_name,
                "annotation_column_used": annotation_column,
                "match_status": status,
            }
        )
        counts[(expression_sample_id, cell_type)] += 1
    total = len(expression_cells)
    match_rate = matched / total if total else 0.0
    warning = "ok" if match_rate >= 0.90 else "low_annotation_match_rate_review_cell_id_normalization"
    summary_rows = [
        {
            "total_expression_cells": str(total),
            "matched_expression_cells": str(matched),
            "unmatched_expression_cells": str(total - matched),
            "unmatched_cells_total": str(total - matched),
            "match_rate": f"{match_rate:.6g}",
            "annotation_rows": str(len(workbook_rows)),
            "selected_celltype_column": annotation_column,
            "chosen_annotation_column": annotation_column,
            "annotation_candidate_columns": ";".join(row["column_name"] for row in candidate_rows if row.get("column_name")),
            "selected_sheet": sheet_name,
            "warning": warning,
        }
    ]
    per_celltype = [
        {
            "sample_id": sample_id,
            "cell_type": cell_type,
            "cell_count": str(count),
            "selected_celltype_column": annotation_column,
            "selected_sheet": sheet_name,
            "match_rate": f"{match_rate:.6g}",
        }
        for (sample_id, cell_type), count in sorted(counts.items())
    ]
    return output_rows, summary_rows, [*candidate_rows, *per_celltype]


def write_tsv(path: Path, rows: list[dict[str, str]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = sorted({field for row in rows for field in row}) if rows else ["status"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build repaired GSE243639 cell annotation map.")
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--output", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_annotation_map.tsv"))
    parser.add_argument("--summary-output", type=Path, default=Path("results/tables/phase18_gse243639_annotation_match_summary.tsv"))
    parser.add_argument("--celltype-summary-output", type=Path, default=Path("results/tables/phase17_gse243639_cell_annotation_summary.tsv"))
    parser.add_argument("--candidate-output", type=Path, default=Path("results/reports/phase18_gse243639_annotation_column_candidates.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/84_build_gse243639_cell_annotation_map.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    rows, summary_rows, mixed_rows = build_annotation_map(args.xlsx, args.cell_sample_map)
    celltype_rows = [row for row in mixed_rows if "cell_count" in row]
    candidate_rows = [row for row in mixed_rows if "column_name" in row]
    write_tsv(
        args.output,
        rows,
        [
            "cell_id_expression",
            "cell_id_annotation_original",
            "normalized_cell_id",
            "barcode_core",
            "sample_id",
            "cell_type",
            "cluster_id",
            "biological_celltype_confidence",
            "annotation_source_sheet",
            "annotation_column_used",
            "match_status",
        ],
    )
    write_tsv(
        args.summary_output,
        summary_rows,
        [
            "total_expression_cells",
            "matched_expression_cells",
            "unmatched_expression_cells",
            "match_rate",
            "annotation_rows",
            "selected_celltype_column",
            "chosen_annotation_column",
            "annotation_candidate_columns",
            "unmatched_cells_total",
            "selected_sheet",
            "warning",
        ],
    )
    write_tsv(args.celltype_summary_output, celltype_rows, ["sample_id", "cell_type", "cell_count", "selected_celltype_column", "selected_sheet", "match_rate"])
    write_tsv(args.candidate_output, candidate_rows, ["column_name", "annotation_kind", "biological_label_hits", "numeric_fraction", "confidence"])
    logging.info("Wrote repaired GSE243639 annotation rows: %d", len(rows))
    print(f"Wrote {args.output}")
    print(f"Wrote {args.summary_output}")
    print(f"Wrote {args.candidate_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
