#!/usr/bin/env python3
"""Deep normalization audit for GSE243639 expression and workbook cell IDs."""

from __future__ import annotations

import argparse
import csv
import gzip
import logging
import re
from pathlib import Path
from typing import Callable, TextIO


CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]
SAFE_OVERLAP_THRESHOLD = 0.95


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def open_text(path: Path) -> TextIO:
    if path.name.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8", newline="")
    return path.open("r", encoding="utf-8", newline="")


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook ID audit. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_id(value: str) -> tuple[str, str]:
    stripped = value.strip().strip("\"'")
    if "_" in stripped:
        sample_id, barcode = stripped.split("_", 1)
        return sample_id, barcode
    return "", stripped


def remove_trailing(value: str) -> str:
    return re.sub(r"([.-])\d+$", "", value.strip())


def barcode_only(value: str) -> str:
    return remove_trailing(split_id(value)[1])


def sample_only(value: str) -> str:
    return split_id(value)[0]


def collapse_punctuation(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value.strip()).upper()


def seurat_sample_barcode_dot(value: str) -> str:
    sample_id, barcode = split_id(value)
    core = remove_trailing(barcode)
    return f"{sample_id}_{core}.1" if sample_id else f"{core}.1"


def seurat_sample_barcode_dash(value: str) -> str:
    sample_id, barcode = split_id(value)
    core = remove_trailing(barcode)
    return f"{sample_id}_{core}-1" if sample_id else f"{core}-1"


def seurat_barcode_dash_sample(value: str) -> str:
    sample_id, barcode = split_id(value)
    core = remove_trailing(barcode)
    return f"{core}-1_{sample_id}" if sample_id else f"{core}-1"


def seurat_barcode_dot_sample(value: str) -> str:
    sample_id, barcode = split_id(value)
    core = remove_trailing(barcode)
    return f"{core}.1_{sample_id}" if sample_id else f"{core}.1"


def read_expression_ids(path: Path) -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()
    with open_text(path) as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id and cell_id not in seen:
                ids.append(cell_id)
                seen.add(cell_id)
    return ids


def read_cell_sample_ids(path: Path) -> list[str]:
    ids: list[str] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id:
                ids.append(cell_id)
    return ids


def role_score(column: str) -> int:
    lowered = column.lower().replace(" ", "_")
    return sum(1 for term in CELL_ID_TERMS if term in lowered)


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        score = sum(role_score(value) for value in row) + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_cell_column(header: list[str]) -> str:
    scored = sorted(((role_score(column), column) for column in header if column), reverse=True)
    return scored[0][1] if scored and scored[0][0] > 0 else (header[0] if header else "")


def read_workbook_ids(path: Path) -> list[str]:
    workbook = load_workbook_readonly(path)
    ids: list[str] = []
    seen: set[str] = set()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        preview = [[clean(value) for value in row] for row in sheet.iter_rows(max_row=40, values_only=True)]
        if not preview:
            continue
        header_index = infer_header_row(preview)
        header = preview[header_index]
        cell_column = choose_cell_column(header)
        column_index = header.index(cell_column) if cell_column in header else 0
        for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
            values = [clean(value) for value in row]
            if column_index < len(values) and values[column_index] and values[column_index] not in seen:
                ids.append(values[column_index])
                seen.add(values[column_index])
        break
    return ids


def transform_set(ids: list[str], transform: Callable[[str], str]) -> dict[str, str]:
    output: dict[str, str] = {}
    for cell_id in ids:
        key = transform(cell_id)
        if key:
            output.setdefault(key, cell_id)
    return output


def audit_rule(
    rule_name: str,
    expression_ids: list[str],
    workbook_ids: list[str],
    expression_transform: Callable[[str], str],
    workbook_transform: Callable[[str], str],
) -> dict[str, str]:
    expression_keys = transform_set(expression_ids, expression_transform)
    workbook_keys = transform_set(workbook_ids, workbook_transform)
    matches = sorted(set(expression_keys) & set(workbook_keys))
    denominator = max(1, min(len(expression_keys), len(workbook_keys)))
    overlap_rate = len(matches) / denominator
    example_key = matches[0] if matches else ""
    return {
        "rule_name": rule_name,
        "expression_unique_ids": str(len(expression_keys)),
        "workbook_unique_ids": str(len(workbook_keys)),
        "overlap_count": str(len(matches)),
        "overlap_rate": f"{overlap_rate:.8g}",
        "example_expression_id": expression_keys.get(example_key, expression_ids[0] if expression_ids else ""),
        "example_workbook_id": workbook_keys.get(example_key, workbook_ids[0] if workbook_ids else ""),
        "example_matched_id": example_key,
        "safe_to_use": str(overlap_rate >= SAFE_OVERLAP_THRESHOLD).lower(),
        "notes": "Requires manual review even when overlap is high.",
    }


def row_order_candidate(expression_ids: list[str], workbook_ids: list[str]) -> dict[str, str]:
    equal_count = len(expression_ids) == len(workbook_ids) and len(expression_ids) > 0
    return {
        "rule_name": "workbook_row_number_mapping_candidate",
        "expression_unique_ids": str(len(set(expression_ids))),
        "workbook_unique_ids": str(len(set(workbook_ids))),
        "overlap_count": "0",
        "overlap_rate": "0",
        "example_expression_id": expression_ids[0] if expression_ids else "",
        "example_workbook_id": workbook_ids[0] if workbook_ids else "",
        "example_matched_id": "",
        "safe_to_use": "false",
        "notes": "Row-order linkage is only a hypothesis here; use script 97 for independent audit. Equal row count: "
        + str(equal_count).lower(),
    }


def count_header_candidate(cell_sample_ids: list[str], workbook_ids: list[str]) -> dict[str, str]:
    equal_count = len(cell_sample_ids) == len(workbook_ids) and len(cell_sample_ids) > 0
    return {
        "rule_name": "count_header_column_order_mapping_candidate",
        "expression_unique_ids": str(len(set(cell_sample_ids))),
        "workbook_unique_ids": str(len(set(workbook_ids))),
        "overlap_count": "0",
        "overlap_rate": "0",
        "example_expression_id": cell_sample_ids[0] if cell_sample_ids else "",
        "example_workbook_id": workbook_ids[0] if workbook_ids else "",
        "example_matched_id": "",
        "safe_to_use": "false",
        "notes": "Column-order linkage must be audited by script 97 before use. Equal row count: "
        + str(equal_count).lower(),
    }


def build_rows(expression: Path, cell_sample_map: Path, xlsx: Path) -> list[dict[str, str]]:
    expression_ids = read_expression_ids(expression)
    cell_sample_ids = read_cell_sample_ids(cell_sample_map)
    workbook_ids = read_workbook_ids(xlsx)
    rules: list[tuple[str, Callable[[str], str], Callable[[str], str]]] = [
        ("raw_id", lambda value: value, lambda value: value),
        ("lowercase", lambda value: value.lower(), lambda value: value.lower()),
        ("remove_quotes", lambda value: value.strip("\"'"), lambda value: value.strip("\"'")),
        ("replace_dash_with_dot", lambda value: value.replace("-", "."), lambda value: value.replace("-", ".")),
        ("replace_dot_with_dash", lambda value: value.replace(".", "-"), lambda value: value.replace(".", "-")),
        ("remove_trailing_dot_or_dash_one", remove_trailing, remove_trailing),
        ("keep_barcode_only", barcode_only, barcode_only),
        ("keep_sample_prefix_only", sample_only, sample_only),
        ("remove_sample_prefix", lambda value: split_id(value)[1], lambda value: split_id(value)[1]),
        ("collapse_punctuation", collapse_punctuation, collapse_punctuation),
        ("seurat_sample_barcode_dot_one", seurat_sample_barcode_dot, seurat_sample_barcode_dot),
        ("seurat_sample_barcode_dash_one", seurat_sample_barcode_dash, seurat_sample_barcode_dash),
        ("seurat_barcode_dash_one_sample", seurat_barcode_dash_sample, seurat_barcode_dash_sample),
        ("seurat_barcode_dot_one_sample", seurat_barcode_dot_sample, seurat_barcode_dot_sample),
    ]
    rows = [audit_rule(name, expression_ids, workbook_ids, left, right) for name, left, right in rules]
    rows.append(row_order_candidate(cell_sample_ids, workbook_ids))
    rows.append(count_header_candidate(cell_sample_ids, workbook_ids))
    return rows


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = [
        "rule_name",
        "expression_unique_ids",
        "workbook_unique_ids",
        "overlap_count",
        "overlap_rate",
        "example_expression_id",
        "example_workbook_id",
        "example_matched_id",
        "safe_to_use",
        "notes",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    safe_rows = [row for row in rows if row["safe_to_use"] == "true"]
    best = max(rows, key=lambda row: float(row["overlap_rate"])) if rows else None
    lines = [
        "# Phase 19 GSE243639 Best Normalization Rule",
        "",
        "This audit tests string-only normalization strategies. Row-order hypotheses are never accepted by this script.",
        "",
        f"- Safe ID-based rules found: {len(safe_rows)}",
    ]
    if best:
        lines.extend(
            [
                f"- Best rule: {best['rule_name']}",
                f"- Best overlap rate: {best['overlap_rate']}",
                f"- Safe to use: {best['safe_to_use']}",
                f"- Notes: {best['notes']}",
            ]
        )
    if not safe_rows:
        lines.append("- Recommendation: do not use workbook annotations by ID unless another audited rule succeeds.")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Deep GSE243639 cell-ID normalization audit.")
    parser.add_argument("--expression", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_sparse_gene_panel_expression.tsv.gz"))
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("results/tables/phase19_gse243639_deep_cell_id_overlap.tsv"))
    parser.add_argument("--best-rule-output", type=Path, default=Path("results/reports/phase19_gse243639_best_normalization_rule.md"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/96_deep_gse243639_cell_id_normalization_audit.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    rows = build_rows(args.expression, args.cell_sample_map, args.xlsx)
    write_tsv(args.output, rows)
    write_markdown(args.best_rule_output, rows)
    logging.info("Wrote deep normalization audit rows: %d", len(rows))
    print(f"Wrote {args.output}")
    print(f"Wrote {args.best_rule_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
