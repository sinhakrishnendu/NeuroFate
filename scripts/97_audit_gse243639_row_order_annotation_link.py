#!/usr/bin/env python3
"""Audit row-order annotation linkage for GSE243639 as a hypothesis only."""

from __future__ import annotations

import argparse
import csv
import gzip
import logging
import re
from collections import Counter
from pathlib import Path
from typing import TextIO


CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]
SAMPLE_TERMS = ["sample", "donor", "subject"]
ANNOTATION_TERMS = ["celltype", "cell_type", "annotation", "cluster", "subcluster", "class", "type"]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def open_text(path: Path) -> TextIO:
    if path.name.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8", newline="")
    return path.open("r", encoding="utf-8", newline="")


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for row-order audit. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def role_score(column: str, terms: list[str]) -> int:
    lowered = column.lower().replace(" ", "_")
    return sum(1 for term in terms if term in lowered)


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        score = (
            sum(role_score(value, CELL_ID_TERMS) for value in row)
            + sum(role_score(value, SAMPLE_TERMS) for value in row)
            + sum(role_score(value, ANNOTATION_TERMS) for value in row)
            + len([value for value in row if value])
        )
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_column(columns: list[str], terms: list[str]) -> str:
    scored = sorted(((role_score(column, terms), column) for column in columns if column), reverse=True)
    return scored[0][1] if scored and scored[0][0] > 0 else ""


def split_sample(cell_id: str) -> str:
    return cell_id.split("_", 1)[0] if "_" in cell_id else ""


def read_cell_sample_rows(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for index, row in enumerate(reader, start=1):
            cell_id = row.get("cell_id", "")
            sample_id = row.get("sample_id", "") or split_sample(cell_id)
            rows.append({"row_index": str(index), "cell_id": cell_id, "sample_id": sample_id})
    return rows


def read_clinical_samples(path: Path, header_line: int = 6, delimiter: str = ";") -> set[str]:
    samples: set[str] = set()
    with open_text(path) as handle:
        for _ in range(header_line - 1):
            next(handle, "")
        reader = csv.DictReader(handle, delimiter=delimiter)
        for row in reader:
            sample_id = (row.get("Sample ID") or "").strip()
            if sample_id:
                samples.add(sample_id)
    return samples


def read_workbook_sheet_rows(path: Path) -> dict[str, dict[str, object]]:
    workbook = load_workbook_readonly(path)
    sheet_data: dict[str, dict[str, object]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        preview = [[clean(value) for value in row] for row in sheet.iter_rows(max_row=40, values_only=True)]
        if not preview:
            sheet_data[sheet_name] = {"columns": [], "rows": [], "cell_column": "", "sample_column": "", "annotation_column": ""}
            continue
        header_index = infer_header_row(preview)
        columns = preview[header_index]
        cell_column = choose_column(columns, CELL_ID_TERMS) or (columns[0] if columns else "")
        sample_column = choose_column(columns, SAMPLE_TERMS)
        annotation_column = choose_column(columns, ANNOTATION_TERMS)
        rows: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
            values = [clean(value) for value in row]
            if any(values):
                rows.append({column: value for column, value in zip(columns, values, strict=False) if column})
        sheet_data[sheet_name] = {
            "columns": columns,
            "rows": rows,
            "cell_column": cell_column,
            "sample_column": sample_column,
            "annotation_column": annotation_column,
        }
    return sheet_data


def run_lengths(values: list[str], limit: int = 20) -> list[str]:
    if not values:
        return []
    runs: list[str] = []
    current = values[0]
    count = 0
    for value in values:
        if value == current:
            count += 1
        else:
            runs.append(f"{current}:{count}")
            current = value
            count = 1
        if len(runs) >= limit:
            break
    if len(runs) < limit:
        runs.append(f"{current}:{count}")
    return runs


def sample_distribution(values: list[str]) -> str:
    counts = Counter(value for value in values if value)
    return ";".join(f"{key}:{value}" for key, value in counts.most_common(10))


def annotation_distribution(rows: list[dict[str, str]], column: str) -> str:
    if not column:
        return "annotation_column_unavailable"
    counts = Counter(row.get(column, "") for row in rows if row.get(column, ""))
    return ";".join(f"{key}:{value}" for key, value in counts.most_common(10))


def first_last_summary(values: list[str], limit: int = 100) -> tuple[str, str]:
    first = sample_distribution(values[:limit])
    last = sample_distribution(values[-limit:]) if values else ""
    return first, last


def decide_row_order(
    count_equal: bool,
    sample_column_present: bool,
    sample_values_match_clinical: bool,
    first_last_consistent: bool,
    sheet_consistent: bool,
) -> tuple[str, str]:
    if count_equal and sample_column_present and sample_values_match_clinical and first_last_consistent and sheet_consistent:
        return "safe_row_order_linkage", "Exact row counts plus independent sample grouping checks support cautious row-order use."
    if not count_equal:
        return "unsafe_row_order_linkage", "Workbook and count-header/cell-map row counts differ."
    if not sample_column_present:
        return "inconclusive_row_order_linkage", "Workbook lacks an independently usable sample/donor column."
    return "inconclusive_row_order_linkage", "Row counts match, but independent consistency checks are incomplete."


def build_audit(cell_sample_map: Path, xlsx: Path, clinical: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    map_rows = read_cell_sample_rows(cell_sample_map)
    clinical_samples = read_clinical_samples(clinical)
    workbook_data = read_workbook_sheet_rows(xlsx)
    primary_sheet = next(iter(workbook_data))
    primary = workbook_data[primary_sheet]
    workbook_rows = primary["rows"]  # type: ignore[index]
    sample_column = str(primary["sample_column"])
    annotation_column = str(primary["annotation_column"])
    map_sample_values = [row["sample_id"] for row in map_rows]
    workbook_sample_values = [row.get(sample_column, "") for row in workbook_rows] if sample_column else []
    count_equal = len(map_rows) == len(workbook_rows) and len(map_rows) > 0
    sample_column_present = bool(sample_column)
    sample_values_match_clinical = bool(workbook_sample_values) and set(workbook_sample_values) <= clinical_samples
    first_map, last_map = first_last_summary(map_sample_values)
    first_book, last_book = first_last_summary(workbook_sample_values)
    first_last_consistent = bool(workbook_sample_values) and first_map == first_book and last_map == last_book
    sheet_row_counts = {sheet: len(data["rows"]) for sheet, data in workbook_data.items()}  # type: ignore[arg-type]
    sheet_consistent = len(set(sheet_row_counts.values())) <= 1
    decision, reason = decide_row_order(
        count_equal,
        sample_column_present,
        sample_values_match_clinical,
        first_last_consistent,
        sheet_consistent,
    )
    audit_rows = [
        {
            "check_name": "row_count_equality",
            "value": str(count_equal).lower(),
            "details": f"cell_sample_map_rows={len(map_rows)}; workbook_rows={len(workbook_rows)}",
            "decision": decision,
            "notes": reason,
        },
        {
            "check_name": "workbook_sample_column_present",
            "value": str(sample_column_present).lower(),
            "details": f"sample_column={sample_column or 'unavailable'}",
            "decision": decision,
            "notes": "Independent sample labels are required before row-order linkage can be trusted.",
        },
        {
            "check_name": "workbook_samples_match_clinical",
            "value": str(sample_values_match_clinical).lower(),
            "details": f"clinical_samples={len(clinical_samples)}; workbook_sample_distribution={sample_distribution(workbook_sample_values)}",
            "decision": decision,
            "notes": "Workbook sample labels must map into clinical sample IDs.",
        },
        {
            "check_name": "first_last_100_sample_grouping_consistent",
            "value": str(first_last_consistent).lower(),
            "details": f"map_first={first_map}; workbook_first={first_book}; map_last={last_map}; workbook_last={last_book}",
            "decision": decision,
            "notes": "This checks order grouping but does not prove cell identity.",
        },
        {
            "check_name": "sheet_row_order_consistency",
            "value": str(sheet_consistent).lower(),
            "details": ";".join(f"{sheet}:{count}" for sheet, count in sheet_row_counts.items()),
            "decision": decision,
            "notes": "Workbook sheets should have consistent row counts for a row-order hypothesis.",
        },
        {
            "check_name": "annotation_distribution_preview",
            "value": annotation_distribution(workbook_rows, annotation_column),
            "details": f"annotation_column={annotation_column or 'unavailable'}",
            "decision": decision,
            "notes": "Distribution is descriptive only and cannot independently prove row-order linkage.",
        },
    ]
    preview_rows: list[dict[str, str]] = []
    for index in list(range(0, min(20, len(map_rows)))) + list(range(max(0, len(map_rows) - 20), len(map_rows))):
        if index >= len(workbook_rows):
            continue
        workbook_row = workbook_rows[index]
        preview_rows.append(
            {
                "row_index": str(index + 1),
                "cell_sample_map_cell_id": map_rows[index]["cell_id"],
                "cell_sample_map_sample_id": map_rows[index]["sample_id"],
                "workbook_cell_id": workbook_row.get(str(primary["cell_column"]), ""),
                "workbook_sample_id": workbook_row.get(sample_column, "") if sample_column else "",
                "workbook_annotation": workbook_row.get(annotation_column, "") if annotation_column else "",
                "row_order_decision": decision,
            }
        )
    return audit_rows, preview_rows


def write_tsv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit GSE243639 row-order annotation linkage without accepting it automatically.")
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--clinical", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_Clinical_data.csv.gz"))
    parser.add_argument("--output", type=Path, default=Path("results/tables/phase19_gse243639_row_order_link_audit.tsv"))
    parser.add_argument("--preview-output", type=Path, default=Path("results/reports/phase19_gse243639_row_order_link_preview.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/97_audit_gse243639_row_order_annotation_link.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    audit_rows, preview_rows = build_audit(args.cell_sample_map, args.xlsx, args.clinical)
    write_tsv(args.output, audit_rows, ["check_name", "value", "details", "decision", "notes"])
    write_tsv(
        args.preview_output,
        preview_rows,
        [
            "row_index",
            "cell_sample_map_cell_id",
            "cell_sample_map_sample_id",
            "workbook_cell_id",
            "workbook_sample_id",
            "workbook_annotation",
            "row_order_decision",
        ],
    )
    logging.info("Row-order linkage decision: %s", audit_rows[0]["decision"] if audit_rows else "unavailable")
    print(f"Wrote {args.output}")
    print(f"Wrote {args.preview_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
