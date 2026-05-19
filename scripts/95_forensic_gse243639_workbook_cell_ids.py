#!/usr/bin/env python3
"""Forensically preview GSE243639 expression, map, and workbook cell IDs."""

from __future__ import annotations

import argparse
import csv
import gzip
import logging
import re
from collections import Counter
from pathlib import Path
from typing import Iterable, TextIO


CELL_ID_TERMS = ["cell", "barcode", "nucleus", "nuclei", "cell_id", "cellid"]


def configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def open_text(path: Path) -> TextIO:
    if path.name.endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8", newline="")
    return path.open("r", encoding="utf-8", newline="")


def load_workbook_readonly(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook ID forensics. Install it manually before running this script.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def remove_trailing_suffix(value: str) -> str:
    return re.sub(r"([.-])\d+$", "", value.strip())


def split_sample_barcode(cell_id: str) -> tuple[str, str]:
    value = cell_id.strip()
    if "_" in value:
        sample_id, barcode = value.split("_", 1)
        return sample_id, barcode
    return "", value


def barcode_core(cell_id: str) -> str:
    _, barcode = split_sample_barcode(cell_id)
    return remove_trailing_suffix(barcode)


def collapse_punctuation(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value).upper()


def delimiter_pattern(value: str) -> str:
    delimiters = [char for char in ["_", "-", ".", ":", "|", "/"] if char in value]
    return "".join(delimiters) if delimiters else "none"


def id_type_guess(value: str) -> str:
    stripped = value.strip()
    if not stripped:
        return "blank"
    if re.fullmatch(r"\d+(\.0)?", stripped):
        return "numeric_id"
    if "_" in stripped and re.search(r"[ACGTN]{8,}", stripped, re.IGNORECASE):
        return "sample_prefixed_barcode"
    if re.search(r"[ACGTN]{8,}", stripped, re.IGNORECASE):
        return "barcode_like"
    if re.search(r"^[A-Za-z]+[_-]\d+", stripped):
        return "seurat_or_object_id"
    return "unrelated_or_custom_id"


def pattern_summary(values: Iterable[str]) -> str:
    guesses = Counter(id_type_guess(value) for value in values if value)
    if not guesses:
        return "no_ids"
    return ";".join(f"{key}:{value}" for key, value in guesses.most_common())


def expression_preview(path: Path, limit: int = 50) -> list[str]:
    ids: list[str] = []
    with open_text(path) as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        seen: set[str] = set()
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id and cell_id not in seen:
                ids.append(cell_id)
                seen.add(cell_id)
            if len(ids) >= limit:
                break
    return ids


def cell_sample_preview(path: Path, limit: int = 50) -> list[str]:
    ids: list[str] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            cell_id = row.get("cell_id", "")
            if cell_id:
                ids.append(cell_id)
            if len(ids) >= limit:
                break
    return ids


def role_score(column: str) -> int:
    lowered = column.lower().replace(" ", "_")
    return sum(1 for term in CELL_ID_TERMS if term in lowered)


def infer_header_row(rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        score = sum(role_score(value) for value in row) + len([value for value in row if value])
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_cell_column(header: list[str]) -> str:
    scored = sorted(((role_score(column), column) for column in header if column), reverse=True)
    return scored[0][1] if scored and scored[0][0] > 0 else (header[0] if header else "")


def workbook_previews(path: Path, limit: int = 50) -> dict[str, tuple[str, list[str]]]:
    workbook = load_workbook_readonly(path)
    previews: dict[str, tuple[str, list[str]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        preview_rows = [[clean(value) for value in row] for row in sheet.iter_rows(max_row=40, values_only=True)]
        if not preview_rows:
            previews[sheet_name] = ("", [])
            continue
        header_index = infer_header_row(preview_rows)
        header = preview_rows[header_index]
        cell_column = choose_cell_column(header)
        try:
            column_index = header.index(cell_column)
        except ValueError:
            column_index = 0
        ids: list[str] = []
        for row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
            values = [clean(value) for value in row]
            if column_index < len(values) and values[column_index]:
                ids.append(values[column_index])
            if len(ids) >= limit:
                break
        previews[sheet_name] = (cell_column, ids)
    return previews


def preview_rows(source: str, sheet_name: str, cell_column: str, ids: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    summary = pattern_summary(ids)
    for index, cell_id in enumerate(ids[:50], start=1):
        sample_id, barcode = split_sample_barcode(cell_id)
        rows.append(
            {
                "source": source,
                "sheet_name": sheet_name,
                "cell_column": cell_column,
                "row_index": str(index),
                "cell_id": cell_id,
                "string_length": str(len(cell_id)),
                "sample_prefix": sample_id,
                "barcode_part": barcode,
                "barcode_core": barcode_core(cell_id),
                "delimiter_pattern": delimiter_pattern(cell_id),
                "normalized_form": collapse_punctuation(remove_trailing_suffix(cell_id)),
                "id_type_guess": id_type_guess(cell_id),
                "source_pattern_summary": summary,
            }
        )
    return rows


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = [
        "source",
        "sheet_name",
        "cell_column",
        "row_index",
        "cell_id",
        "string_length",
        "sample_prefix",
        "barcode_part",
        "barcode_core",
        "delimiter_pattern",
        "normalized_form",
        "id_type_guess",
        "source_pattern_summary",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Preview GSE243639 workbook cell IDs without linking annotations.")
    parser.add_argument("--expression", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_sparse_gene_panel_expression.tsv.gz"))
    parser.add_argument("--cell-sample-map", type=Path, default=Path("data/interim/external/gse243639_pd_snpc/gse243639_cell_sample_map.tsv"))
    parser.add_argument("--xlsx", type=Path, default=Path("data/raw/external/gse243639_pd_snpc/GSE243639_UMAP_coordinates.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("results/reports/phase19_gse243639_cell_id_forensic_preview.tsv"))
    parser.add_argument("--log-file", type=Path, default=Path("results/logs/95_forensic_gse243639_workbook_cell_ids.log"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.log_file)
    rows: list[dict[str, str]] = []
    rows.extend(preview_rows("expression_sparse_table", "", "cell_id", expression_preview(args.expression)))
    rows.extend(preview_rows("cell_sample_map", "", "cell_id", cell_sample_preview(args.cell_sample_map)))
    for sheet_name, (cell_column, ids) in workbook_previews(args.xlsx).items():
        rows.extend(preview_rows("workbook", sheet_name, cell_column, ids))
    write_tsv(args.output, rows)
    logging.info("Wrote forensic preview rows: %d", len(rows))
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
